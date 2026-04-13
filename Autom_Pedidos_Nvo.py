import pandas as pd
import os
from openpyxl import load_workbook

def calcular_pedido_y_generar_excel():
    # Definir rutas de archivos
    ruta_base = "/Users/alejandroyankilevich/Documents/MASTER DATA SCIENCE/Clases/GRIDO/PEDIDOS"
    archivo_stock = os.path.join(ruta_base, "Stock.xlsx")  # Corregido: Stock.xlsx con mayúscula
    archivo_venta = os.path.join(ruta_base, "Ventas_Estimadas.xlsx")
    archivo_modelo = os.path.join(ruta_base, "Modelo_Carrito_Actual.xlsx")
    # Ya no necesitamos archivo_salida porque actualizaremos directamente el modelo

    # Leer archivos de stock y venta estimada
    try:
        df_stock = pd.read_excel(archivo_stock)
        df_venta = pd.read_excel(archivo_venta)
        print(f"✅ Archivos leídos correctamente:")
        print(f"   - Stock: {archivo_stock}")
        print(f"   - Venta: {archivo_venta}")
    except Exception as e:
        print(f"❌ Error al leer archivos: {e}")
        return

    # Mostrar columnas disponibles para debugging
    print(f"\n📋 Columnas en archivo de stock: {list(df_stock.columns)}")
    print(f"📋 Columnas en archivo de venta: {list(df_venta.columns)}")

    # Unir por columna de producto (ajustar nombre si es necesario)
    # Suponemos que ambos tienen columna 'Producto'
    if 'Producto' not in df_stock.columns or 'Producto' not in df_venta.columns:
        print("❌ No se encuentra la columna 'Producto' en uno de los archivos.")
        return

    # Renombrar columnas antes del merge para evitar conflictos
    df_venta_renamed = df_venta.rename(columns={'Venta Estimada': 'Venta_Estimada'})
    df_stock_renamed = df_stock.rename(columns={'Stock Actual': 'Stock_Actual'})
    
    # Hacer el merge con las columnas renombradas
    df = pd.merge(df_venta_renamed, df_stock_renamed, on='Producto', how='left')
    
    # Mostrar las columnas después del merge para debugging
    print(f"\n🔗 Merge realizado. Filas resultantes: {len(df)}")
    print(f"📋 Columnas después del merge: {list(df.columns)}")
    
    # Calcular reposición: Ventas Estimadas - Stock Actual
    if 'Venta_Estimada' in df.columns and 'Stock_Actual' in df.columns:
        print(f"✅ Columnas encontradas: Venta_Estimada y Stock_Actual")
        
        # Calcular reposición: Ventas Estimadas - Stock Actual
        df['A_PEDIR'] = df['Venta_Estimada'] - df['Stock_Actual']
        df['A_PEDIR'] = df['A_PEDIR'].apply(lambda x: max(x, 0) if pd.notna(x) else 0)
        
        # Mostrar resumen del cálculo
        print(f"\n📊 Resumen del cálculo de reposición:")
        print(f"   - Productos procesados: {len(df)}")
        print(f"   - Total a pedir: {df['A_PEDIR'].sum():.0f}")
        print(f"   - Productos con pedido > 0: {(df['A_PEDIR'] > 0).sum()}")
        
        # Mostrar algunos ejemplos del cálculo
        print(f"\n🔍 Ejemplos del cálculo:")
        for i, row in df.head(5).iterrows():
            venta = row['Venta_Estimada']
            stock = row['Stock_Actual']
            pedir = row['A_PEDIR']
            print(f"   {row['Producto']}: {venta} (venta) - {stock} (stock) = {pedir} (a pedir)")
            
    else:
        print(f"❌ Error: Columnas no encontradas después del merge")
        print(f"   Columnas disponibles: {list(df.columns)}")
        return

    # Leer formato del archivo "Modelo Carrito"
    try:
        df_modelo = pd.read_excel(archivo_modelo)
        print(f"✅ Modelo carrito leído: {archivo_modelo}")
    except Exception as e:
        print(f"❌ Error al leer el archivo Modelo Carrito: {e}")
        return

    # Actualizar solo la columna C del archivo Excel manteniendo todo el formato
    print(f"📋 Actualizando solo la columna C del archivo Excel...")
    
    # Cargar el archivo Excel con openpyxl para preservar formato
    try:
        workbook = load_workbook(archivo_modelo)
        worksheet = workbook.active
        print(f"✅ Archivo Excel cargado: {archivo_modelo}")
        
        # Mapear productos y actualizar solo la columna C
        productos_mapeados = 0
        fila = 2  # Empezar desde la fila 2 (asumiendo que la fila 1 tiene encabezados)
        
        # Buscar la columna que contiene la descripción del producto (no el código)
        # Asumimos que puede estar en la columna A, B o C, buscamos la que tenga texto descriptivo
        col_descripcion = None
        for col in [1, 2, 3]:  # Columnas A, B, C
            if worksheet.cell(row=1, column=col).value:
                header = str(worksheet.cell(row=1, column=col).value).lower()
                if any(palabra in header for palabra in ['descripcion', 'producto', 'nombre', 'item']):
                    col_descripcion = col
                    break
        
        # Si no encontramos encabezado descriptivo, usar la columna B (asumiendo que A es código y B es descripción)
        if col_descripcion is None:
            col_descripcion = 2  # Columna B
        
        print(f"🔍 Usando columna {chr(64 + col_descripcion)} para buscar productos por descripción")
        
        while worksheet.cell(row=fila, column=col_descripcion).value is not None:
            producto_modelo = worksheet.cell(row=fila, column=col_descripcion).value
            
            if producto_modelo:
                # Buscar el producto en nuestro cálculo (búsqueda más flexible)
                producto_encontrado = None
                
                # Primero buscar coincidencia exacta
                producto_encontrado = df[df['Producto'] == producto_modelo]
                
                # Si no hay coincidencia exacta, buscar coincidencia parcial
                if producto_encontrado.empty:
                    # Buscar productos que contengan la descripción del modelo
                    for idx, row in df.iterrows():
                        if str(producto_modelo).lower() in str(row['Producto']).lower():
                            producto_encontrado = df.iloc[[idx]]
                            break
                
                # Si aún no hay coincidencia, buscar productos que contengan palabras clave
                if producto_encontrado.empty:
                    palabras_clave = str(producto_modelo).split()
                    for palabra in palabras_clave:
                        if len(palabra) > 3:  # Solo palabras de más de 3 caracteres
                            for idx, row in df.iterrows():
                                if palabra.lower() in str(row['Producto']).lower():
                                    producto_encontrado = df.iloc[[idx]]
                                    break
                            if not producto_encontrado.empty:
                                break
                
                if not producto_encontrado.empty:
                    # Si encontramos el producto, actualizar solo la columna C
                    cantidad_a_pedir = producto_encontrado.iloc[0]['A_PEDIR']
                    worksheet.cell(row=fila, column=3, value=cantidad_a_pedir)  # Columna C = 3
                    print(f"   ✅ {producto_modelo}: {cantidad_a_pedir} unidades → Columna C, Fila {fila}")
                    print(f"      🔍 Coincidencia encontrada con: {producto_encontrado.iloc[0]['Producto']}")
                    productos_mapeados += 1
                else:
                    # Si no encontramos el producto, mantener el valor original
                    print(f"   ⚠️  {producto_modelo}: No encontrado en ventas/stock - manteniendo valor original")
            
            fila += 1
        
        print(f"\n📊 Resumen del mapeo:")
        print(f"   - Filas procesadas: {fila - 2}")
        print(f"   - Productos mapeados exitosamente: {productos_mapeados}")
        print(f"   - Productos con pedido > 0: {len(df[df['A_PEDIR'] > 0])}")
        
    except Exception as e:
        print(f"❌ Error al cargar el archivo Excel: {e}")
        return

    # Guardar el archivo Excel manteniendo todo el formato
    try:
        workbook.save(archivo_modelo)
        print(f"\n✅ Archivo Excel actualizado exitosamente:")
        print(f"   📁 Ubicación: {archivo_modelo}")
        print(f"   🔄 Solo se actualizó la columna C con las cantidades de reposición")
        print(f"   💡 Se preservó todo el formato original (colores, fuentes, fórmulas, etc.)")
        print(f"   📊 Cantidades calculadas: Ventas_Estimadas - Stock_Actual")
    except Exception as e:
        print(f"❌ Error al guardar el archivo Excel: {e}")
        print(f"   💡 Verifica que el archivo no esté abierto en Excel")

if __name__ == "__main__":
    calcular_pedido_y_generar_excel()
